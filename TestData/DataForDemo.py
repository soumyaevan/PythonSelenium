import openpyxl


class DataForDemo:
    data = [{"username": "sensoumya94@gmail.com", "password": "abcd123"},{"username": "sensoumya94@gmail.com", "password": "123abcd"}]

    @staticmethod
    def getDataForARow(testName):
        book = openpyxl.load_workbook('/Users/soumyasen/PycharmProjects/SeleniumTest/TestData/Data.xlsx')
        sheet = book.active
        maxRow = sheet.max_row
        maxCol = sheet.max_column
        data =[]
        Dict = {}
        for r in range(1, maxRow + 1):
            if (sheet.cell(row=r, column=1).value == testName):
                for c in range(2, maxCol + 1):
                    Dict[sheet.cell(row=1, column=c).value] = sheet.cell(row=r, column=c).value
                data.append(Dict.copy())
        return data