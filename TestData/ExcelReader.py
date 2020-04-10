import openpyxl


def readData(rowNamwe, colName):
    book = openpyxl.load_workbook('Data.xlsx')
    sheet = book.active
    maxRow = sheet.max_row
    maxCol = sheet.max_column
    rowVal = None
    colVal = None
    for r in range(1,maxRow+1):
        if (sheet.cell(row=r, column=1).value == rowNamwe):
            rowVal = r
            for c in range(1,maxCol+1):
                if (sheet.cell(row=1, column=c).value == colName):
                    colVal = c
                    return sheet.cell(row=rowVal, column=colVal).value

def getDataForARow(testName):
    book = openpyxl.load_workbook('Data.xlsx')
    sheet = book.active
    maxRow = sheet.max_row
    maxCol = sheet.max_column
    Dict = {}
    for r in range(1, maxRow + 1):
        if (sheet.cell(row=r, column=1).value == testName):
            for c in range(2, maxCol + 1):
                Dict[sheet.cell(row=1, column=c).value] = sheet.cell(row=r, column=c).value
    return Dict

# print(getDataForARow('TestCase2'))

for k,v in getDataForARow('TestCase2').items():
    print(k + " : " + v)

# print(readData('TestCase2','username'))